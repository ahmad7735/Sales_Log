import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import time
import io
import os
import tempfile


# File path
EXCEL_FILE = "data.xlsx"
import os

# ---------------- Data IO ----------------
# Removing @st.cache_data to avoid caching issues
def load_data():
    sales = pd.read_excel(EXCEL_FILE, sheet_name="SalesLog")
    collections = pd.read_excel(EXCEL_FILE, sheet_name="Collections")
    assignments = pd.read_excel(EXCEL_FILE, sheet_name="Assignments")

    # Ensure numeric
    sales["QuotedPrice"] = pd.to_numeric(sales.get("QuotedPrice", 0), errors="coerce").fillna(0)
    sales["DepositPaid"] = pd.to_numeric(sales.get("DepositPaid", 0), errors="coerce").fillna(0)

    # Ensure collections fields are numeric
    collections["QuoteID"] = pd.to_numeric(collections.get("QuoteID", 0), errors="coerce")
    collections["DepositPaid"] = pd.to_numeric(collections.get("DepositPaid", 0), errors="coerce").fillna(0)

    # Ensure dates
    if "StartDate" in sales.columns:
        sales["StartDate"] = pd.to_datetime(sales["StartDate"], errors="coerce")
    if "EndDate" in sales.columns:
        sales["EndDate"] = pd.to_datetime(sales["EndDate"], errors="coerce")

    # Ensure QuoteID column exists
    if "QuoteID" not in sales.columns:
        sales["QuoteID"] = pd.Series(dtype="int")

    # Normalize QuoteID
    sales["QuoteID"] = pd.to_numeric(sales["QuoteID"], errors="coerce").fillna(0).astype(int)
    collections["QuoteID"] = pd.to_numeric(collections["QuoteID"], errors="coerce").fillna(0).astype(int)

    # 🚫 Drop DepositDue if still present in file
    collections = collections.drop(columns=["DepositDue"], errors="ignore")

    # Ensure Status exists
    if "Status" not in collections.columns:
        collections["Status"] = ""

    return sales, collections, assignments


def save_data(sales, collections, assignments):
    SALES_ORDER = ["QuoteID", "Client", "QuotedPrice", "Status", "SalesRep",
                   "Deposit%", "DepositPaid", "StartDate", "EndDate", "JobType"]
    # No DepositDue here ✅
    COLLECTIONS_ORDER = ["QuoteID", "Client", "DepositPaid", "BalanceDue", "Status"]

    # Normalize key columns
    for df in (sales, collections, assignments):
        if "QuoteID" in df.columns:
            df["QuoteID"] = pd.to_numeric(df["QuoteID"], errors="coerce").fillna(0).astype(int)

    def ensure_and_order(df, desired, fill_defaults=None):
        df = df.copy()
        fill_defaults = fill_defaults or {}
        # make sure desired cols exist
        for col in desired:
            if col not in df.columns:
                df[col] = fill_defaults.get(col, pd.NA)
        # drop DepositDue explicitly if present
        if "DepositDue" in df.columns:
            df = df.drop(columns=["DepositDue"])
        # order: desired first, then extras
        extras = [c for c in df.columns if c not in desired]
        return df[desired + extras]

    sales_to_save = ensure_and_order(
        sales, SALES_ORDER,
        fill_defaults={"Deposit%": 0.0, "DepositPaid": 0.0}
    )
    collections_to_save = ensure_and_order(
        collections, COLLECTIONS_ORDER,
        fill_defaults={"DepositPaid": 0.0, "BalanceDue": 0.0, "Status": ""}
    )
    assignments_to_save = assignments.copy()

    dir_name = os.path.dirname(os.path.abspath(EXCEL_FILE))
    base_name = os.path.basename(EXCEL_FILE)
    tmp_path = None

    try:
        with tempfile.NamedTemporaryFile(delete=False, dir=dir_name, prefix=base_name, suffix=".xlsx") as tmp:
            tmp_path = tmp.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="w") as writer:
            sales_to_save.to_excel(writer, sheet_name="SalesLog", index=False)
            collections_to_save.to_excel(writer, sheet_name="Collections", index=False)
            assignments_to_save.to_excel(writer, sheet_name="Assignments", index=False)

            # >>> Apply Excel number formats (currency + date + percent display) <<<
            from openpyxl.utils import get_column_letter

            wb = writer.book
            ws_sales = writer.sheets["SalesLog"]
            ws_col = writer.sheets["Collections"]

            currency_fmt = '"$"#,##0.00'
            date_fmt = 'yyyy-mm-dd'
            percent_literal_fmt = '0.00"%"'  # shows 20 as 20.00% (no scaling)

            def format_column(ws, df, col_name, num_fmt):
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1  # 1-based
                    col_letter = get_column_letter(col_idx)
                    # skip header at row 1
                    for cell in ws[col_letter][1:]:
                        cell.number_format = num_fmt

            # SalesLog: $ for money, date-only for dates, literal % for Deposit%
            format_column(ws_sales, sales_to_save, "QuotedPrice", currency_fmt)
            format_column(ws_sales, sales_to_save, "DepositPaid", currency_fmt)
            format_column(ws_sales, sales_to_save, "StartDate", date_fmt)
            format_column(ws_sales, sales_to_save, "EndDate", date_fmt)
            format_column(ws_sales, sales_to_save, "Deposit%", percent_literal_fmt)

            # Collections: $ for money
            format_column(ws_col, collections_to_save, "DepositPaid", currency_fmt)
            format_column(ws_col, collections_to_save, "BalanceDue", currency_fmt)

        os.replace(tmp_path, os.path.abspath(EXCEL_FILE))  # atomic replace
        return True
    except PermissionError:
        st.error("⚠️ Could not save data. Please close 'data.xlsx' (Excel might have it open) and try again.")
    except Exception as e:
        st.error(f"💥 Save failed: {type(e).__name__}: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except Exception: pass
    return False



# ---------------- Derivations ----------------
def sync_deposit_paid(sales: pd.DataFrame, collections: pd.DataFrame) -> pd.DataFrame:
    """
    Make Sales.DepositPaid = SUM(Collections.DepositPaid) per QuoteID.
    This avoids double counting and keeps Sales as the single source of truth for totals.
    """
    s = sales.copy()
    c = collections.copy()

    # Normalize
    for df in (s, c):
        if "QuoteID" in df.columns:
            df["QuoteID"] = pd.to_numeric(df["QuoteID"], errors="coerce").fillna(0).astype(int)

    s["QuotedPrice"]   = pd.to_numeric(s.get("QuotedPrice", 0), errors="coerce").fillna(0.0)
    c["DepositPaid"]   = pd.to_numeric(c.get("DepositPaid", 0), errors="coerce").fillna(0.0)

    # Sum every collection (initial + follow-ups) per QuoteID
    sums = c.groupby("QuoteID", dropna=False)["DepositPaid"].sum(min_count=1)
    s["DepositPaid"] = s["QuoteID"].map(sums).fillna(0.0).astype(float)

    # Recompute %
    s["Deposit%"] = s.apply(
        lambda r: round((r["DepositPaid"] / r["QuotedPrice"]) * 100, 2) if r["QuotedPrice"] > 0 else 0.0,
        axis=1
    )
    return s


    # (unused legacy code below intentionally preserved but never reached)
    merged = c.merge(
        s[["QuoteID", "DepositPaid"]].rename(columns={"DepositPaid": "InitialDeposit"}),
        on="QuoteID", how="left"
    )
    merged["InitialDeposit"] = pd.to_numeric(merged["InitialDeposit"], errors="coerce").fillna(0.0)
    raw_sum = merged.groupby("QuoteID")["DepositPaid"].sum()
    has_legacy = (
        (merged["DepositPaid"].round(2) == merged["InitialDeposit"].round(2))
        .groupby(merged["QuoteID"])
        .any()
    )
    initial_map = s.set_index("QuoteID")["DepositPaid"]
    adj_sum = raw_sum - initial_map.where(has_legacy, 0.0)
    adj_sum = adj_sum.reindex(s["QuoteID"].unique(), fill_value=0.0)
    total_paid = initial_map.add(adj_sum, fill_value=0.0)
    s = s.set_index("QuoteID")
    s.loc[total_paid.index, "DepositPaid"] = total_paid.values
    s["Deposit%"] = s.apply(
        lambda r: round((r["DepositPaid"] / r["QuotedPrice"]) * 100, 2) if r["QuotedPrice"] > 0 else 0.0,
        axis=1
    )
    return s.reset_index()



def update_balance_due(sales: pd.DataFrame, collections: pd.DataFrame) -> pd.DataFrame:
    """
    BalanceDue per QuoteID = QuotedPrice - Sales.DepositPaid  (Sales.DepositPaid is TOTAL to date).
    """
    s = sales.copy()
    c = collections.copy()

    # Normalize
    for df in (s, c):
        if "QuoteID" in df.columns:
            df["QuoteID"] = pd.to_numeric(df["QuoteID"], errors="coerce").fillna(0).astype(int)
    s["QuotedPrice"] = pd.to_numeric(s.get("QuotedPrice", 0), errors="coerce").fillna(0.0)
    s["DepositPaid"] = pd.to_numeric(s.get("DepositPaid", 0), errors="coerce").fillna(0.0)

    price_map = s.set_index("QuoteID")["QuotedPrice"]
    paid_map  = s.set_index("QuoteID")["DepositPaid"]
    balance   = (price_map - paid_map).clip(lower=0.0)

    if not c.empty:
        c["BalanceDue"] = c["QuoteID"].map(balance).fillna(0.0).astype(float)
    else:
        if "BalanceDue" not in c.columns:
            c["BalanceDue"] = pd.Series(dtype="float")
    return c



# ---------------- Utils ----------------
# UI display formatting helpers (keep numbers numeric in memory)
def _fmt_currency_series(s):
    return s.apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else "")

def _fmt_percent_series(s):  # NEW: show 20 -> "20.00%"
    return s.apply(lambda x: f"{float(x):.2f}%" if pd.notnull(x) else "")

# Generate unique QuoteID based on area (Toms River or Manahawkin)
def generate_unique_quote_id(area, sales):
    area = (area or "").strip()
    s = sales.copy()
    s["QuoteID"] = pd.to_numeric(s["QuoteID"], errors="coerce").fillna(0).astype(int)

    if area == "Toms River":
        candidates = s.loc[(s["QuoteID"] >= 1000) & (s["QuoteID"] < 2000), "QuoteID"]
        return (candidates.max() + 1) if not candidates.empty else 1000

    if area == "Manahawkin":
        candidates = s.loc[(s["QuoteID"] >= 2000) & (s["QuoteID"] < 3000), "QuoteID"]
        return (candidates.max() + 1) if not candidates.empty else 2000

    # Fallback range (optional): 3000+
    candidates = s.loc[s["QuoteID"] >= 3000, "QuoteID"]
    return (candidates.max() + 1) if not candidates.empty else 3000


def safe_rerun():
    try:
        st.rerun()
    except Exception:
        st.session_state["force_rerun"] = not st.session_state.get("force_rerun", False)
        st.stop()

# ---------------- App boot ----------------
# Load data initially
sales, collections, assignments = load_data()

# 👇 add these
sales = sync_deposit_paid(sales, collections)        # roll up initial + collections into Sales.DepositPaid + Deposit%
collections = update_balance_due(sales, collections) # set BalanceDue from Sales totals


# Sidebar
st.sidebar.title("Navigation")
page = st.sidebar.radio("Select Page", ["Dashboard", "Sales Log", "Collections", "Assignments", "View Reports"])

# ---------------- Dashboard ----------------
if page == "Dashboard":
    st.title("📊 Dashboard")

    # ---- Filters ----
    st.sidebar.subheader("Filters")

    # Sales Rep Filter
    reps = ["All"]
    if "SalesRep" in sales.columns:
        reps += sorted(sales["SalesRep"].dropna().unique().tolist())
    selected_rep = st.sidebar.selectbox("Select Sales Rep", reps)

    # Date Range Filter
    if "StartDate" in sales.columns and pd.api.types.is_datetime64_any_dtype(sales["StartDate"]):
        valid_dates = sales["StartDate"].dropna()
        if not valid_dates.empty:
            min_date = valid_dates.min().date()
            max_date = valid_dates.max().date()
            date_range = st.sidebar.date_input("Select Date Range", [min_date, max_date])
            if isinstance(date_range, list) and len(date_range) == 2:
                start_date, end_date = date_range
            else:
                start_date, end_date = min_date, max_date
        else:
            start_date, end_date = None, None
    else:
        start_date, end_date = None, None

    # Apply Filters
    filtered_sales = sales.copy()
    if selected_rep != "All" and "SalesRep" in filtered_sales.columns:
        filtered_sales = filtered_sales[filtered_sales["SalesRep"] == selected_rep]
    if start_date and end_date and "StartDate" in filtered_sales.columns:
        filtered_sales = filtered_sales[
            (filtered_sales["StartDate"] >= pd.to_datetime(start_date)) &
            (filtered_sales["StartDate"] <= pd.to_datetime(end_date))
        ]

    # ---- Won-only view (copy to avoid SettingWithCopy warnings) ----
    if "Status" in filtered_sales.columns:
        won_sales = filtered_sales[filtered_sales["Status"] == "Won"].copy()
    else:
        won_sales = filtered_sales.iloc[0:0].copy()

    # Ensure numeric QuoteID for joins later
    if "QuoteID" in won_sales.columns:
        won_sales["QuoteID"] = pd.to_numeric(won_sales["QuoteID"], errors="coerce").fillna(0).astype(int)
    if "QuoteID" in collections.columns:
        filtered_collections = collections[collections["QuoteID"].isin(
            won_sales["QuoteID"] if "QuoteID" in won_sales.columns else []
        )].copy()
    else:
        filtered_collections = collections.iloc[0:0].copy()

    # ---- Totals / KPIs ----
 # ---- Totals / KPIs ----

    # Safety: (re)build filtered_sales if it's not in scope
    if 'filtered_sales' not in locals():
        filtered_sales = sales.copy()
        # Reapply filters if those vars exist
        if 'selected_rep' in locals() and selected_rep != "All" and "SalesRep" in filtered_sales.columns:
            filtered_sales = filtered_sales[filtered_sales["SalesRep"] == selected_rep]
        if 'start_date' in locals() and 'end_date' in locals() and start_date and end_date and "StartDate" in filtered_sales.columns:
            filtered_sales = filtered_sales[
                (filtered_sales["StartDate"] >= pd.to_datetime(start_date)) &
                (filtered_sales["StartDate"] <= pd.to_datetime(end_date))
            ]

    # Ensure won_sales and filtered_collections exist
    won_sales = (
        filtered_sales[filtered_sales["Status"] == "Won"].copy()
        if "Status" in filtered_sales.columns else filtered_sales.iloc[0:0].copy()
    )
    if "QuoteID" in collections.columns and "QuoteID" in won_sales.columns:
        won_ids = won_sales["QuoteID"].dropna().astype(int)
        filtered_collections = collections[collections["QuoteID"].isin(won_ids)].copy()
    else:
        filtered_collections = collections.iloc[0:0].copy()

    # Core KPIs
    total_revenue_won = float(won_sales["QuotedPrice"].sum()) if "QuotedPrice" in won_sales.columns else 0.0

    # Split collections into: initial deposit (first ledger row per QuoteID) vs follow-ups
    if not filtered_collections.empty:
        filtered_collections["DepositPaid"] = pd.to_numeric(
            filtered_collections.get("DepositPaid", 0), errors="coerce"
        ).fillna(0.0)
        initial_by_q = (
            filtered_collections.reset_index()
            .sort_values(["QuoteID", "index"])
            .groupby("QuoteID")["DepositPaid"]
            .first()
        )
        total_deposit_won = float(initial_by_q.sum())  # initial deposits only
        total_collections_won = float(filtered_collections["DepositPaid"].sum() - total_deposit_won)  # follow-ups only
    else:
        total_deposit_won = 0.0
        total_collections_won = 0.0

    # Balance due = price - TOTAL paid (Sales.DepositPaid already equals ledger sum via sync_deposit_paid)
    balance_due_won_jobs = float(
        (won_sales["QuotedPrice"] - won_sales["DepositPaid"]).clip(lower=0).sum()
    ) if {"QuotedPrice","DepositPaid"}.issubset(won_sales.columns) else 0.0

    # Top metrics
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Quotes Sent", len(filtered_sales))
    col2.metric("Jobs Won", len(won_sales))
    col3.metric("Jobs Pending", len(filtered_sales[filtered_sales["Status"] == "Sent"]) if "Status" in filtered_sales.columns else 0)
    col4.metric("Jobs Lost", len(filtered_sales[filtered_sales["Status"] == "Lost"]) if "Status" in filtered_sales.columns else 0)

    col5, col6, col7, col8 = st.columns(4)
    win_rate = (len(won_sales) / len(filtered_sales) * 100) if len(filtered_sales) > 0 else 0.0
    col5.metric("Win Rate %", f"{win_rate:.1f}%")
    col6.metric("Closed $", f"${total_revenue_won:,.0f}")
    col7.metric("Deposits Paid", f"${total_deposit_won:,.0f}")   # initial deposits only
    col8.metric("Balance Due", f"${balance_due_won_jobs:,.0f}")

    # Follow-ups only
    st.metric("Total Collected (Collections)", f"${total_collections_won:,.0f}")

    # ---- Revenue Breakdown (unchanged logic) ----
    st.subheader("Revenue Breakdown")
    if "Status" in filtered_sales.columns and "QuotedPrice" in filtered_sales.columns:
        revenue_by_status = filtered_sales.groupby("Status")["QuotedPrice"].sum().sort_index()
        total_revenue = float(revenue_by_status.sum())
        if total_revenue > 0:
            fig, ax = plt.subplots()
            bars = ax.bar(
                revenue_by_status.index,
                revenue_by_status.values,
                color=["green" if s == "Won" else "orange" if s == "Lost" else "blue" for s in revenue_by_status.index],
            )
            pct = (revenue_by_status / total_revenue * 100).values
            for bar, p in zip(bars, pct):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height(), f"{p:.1f}%", ha="center", va="bottom")
            ax.set_ylabel("Revenue ($)")
            ax.set_title("Revenue Breakdown by Status")
            st.pyplot(fig)
        else:
            st.warning("No revenue data available for the selected filters.")
    else:
        st.warning("Sales data missing required columns for chart.")

    # ---- Payments Overview (corrected buckets) ----
    st.subheader("Payments Overview")
    if (total_deposit_won > 0) or (total_collections_won > 0) or (balance_due_won_jobs > 0):
        payments_data = {
            "Deposits Paid": total_deposit_won,      # initial only
            "Collections": total_collections_won,    # follow-ups only
            "Balance Due": balance_due_won_jobs,
        }
        total_payments = sum(payments_data.values())
        if total_payments > 0:
            fig2, ax2 = plt.subplots()
            pd.Series(payments_data).plot(kind="bar", ax=ax2, color=["blue", "green", "purple"])
            pct_labels = {k: (v / total_payments * 100) for k, v in payments_data.items()}
            ax2.set_title("Payments Tracking (Won Jobs Only)")
            ax2.set_xticklabels([f"{k} ({v:.1f}%)" for k, v in pct_labels.items()], rotation=45)
            st.pyplot(fig2)
        else:
            st.warning("No payment data available for the selected filters.")
    else:
        st.warning("No payment data available for the selected filters.")

    # ---- Assigned vs Pending Tasks (Won-only) ----
    st.subheader("Task Overview")
    assigned_tasks = assignments["QuoteID"].dropna().astype(int).unique() if "QuoteID" in assignments.columns else []
    won_quote_ids = won_sales["QuoteID"].dropna().astype(int).unique() if "QuoteID" in won_sales.columns else []
    pending_tasks = [qid for qid in won_quote_ids if qid not in set(assigned_tasks)]

    col9, col10 = st.columns(2)
    col9.metric("Assigned Tasks", len(assigned_tasks))
    col10.metric("Pending Tasks", len(pending_tasks))

    fig3, ax3 = plt.subplots()
    ax3.bar(["Assigned", "Pending"], [len(assigned_tasks), len(pending_tasks)], color=["green", "red"])
    ax3.set_title("Assigned vs Pending Tasks (Won Jobs Only)")
    ax3.set_ylabel("Number of Tasks")
    st.pyplot(fig3)

# ---------------- End Dashboard ----------------

# ---------------- Sales Log ----------------
if page == "Sales Log":
    st.title("📝 Sales Log")

    # Initialize session state inputs if not present
    if "quoted_price_input" not in st.session_state:
        st.session_state.quoted_price_input = 0.0
    if "deposit_paid_input" not in st.session_state:
        st.session_state.deposit_paid_input = 0.0
    if "sale_added" not in st.session_state:
        st.session_state["sale_added"] = None

    # Show success message and ask if want to add another sale
    if st.session_state["sale_added"] is not None:
        new_row = st.session_state["sale_added"]
        st.success(f"✅ Sale added successfully! (Quote ID: {new_row['QuoteID']})")

        add_another = st.radio("Would you like to add another sale?", ["No", "Yes"], index=0)

        if add_another == "Yes":
            # Clear session to show form again immediately
            st.session_state["sale_added"] = None
            st.rerun()
        else:
            st.subheader("Latest Added Sale")
            st.dataframe(pd.DataFrame([new_row]))

    # If no sale added or user selected "Yes" to add another
    if st.session_state.get("sale_added") is None:
        quoted_price = st.number_input("Quoted Price *", min_value=0.0, format="%.2f", key="quoted_price_input")
        deposit_paid = st.number_input("Deposit Paid *", min_value=0.0, format="%.2f", key="deposit_paid_input")

        if quoted_price > 0:
            deposit_pct = round((deposit_paid / quoted_price) * 100, 2)
        else:
            deposit_pct = 0.0

        st.text(f"Deposit %: {deposit_pct:.2f}%")

        with st.form("add_sale"):
            st.subheader("Add New Sale")
            area = st.selectbox("Area *", ["Toms River", "Manahawkin"])
            client = st.text_input("Client *")
            status = st.selectbox("Status *", ["Sent", "Won", "Lost"])
            sales_rep = st.text_input("Sales Rep *")
            start_date = st.date_input("Start Date *")
            end_date = st.date_input("End Date *")
            job_type = st.text_input("Job Type *")

            submitted = st.form_submit_button("Add Sale")

            if submitted:
                quoted_price_val = st.session_state.get("quoted_price_input", 0.0)
                deposit_paid_val = st.session_state.get("deposit_paid_input", 0.0)

                if not client or not sales_rep or not job_type:
                    st.error("⚠️ All fields are required. Please fill them in.")
                else:
                    deposit_pct = round((deposit_paid_val / quoted_price_val) * 100, 2) if quoted_price_val > 0 else 0.0

                    # 1) Generate QuoteID
                    new_id = generate_unique_quote_id(area, sales)

                    # 2) Create the Sales row (in-memory)
                    new_row = {
                        "QuoteID": new_id,
                        "Client": client,
                        "QuotedPrice": quoted_price_val,
                        "Status": status,
                        "SalesRep": sales_rep,
                        "Deposit%": deposit_pct,
                        "DepositPaid": deposit_paid_val,   # initial value; sync may overwrite from Collections
                        "StartDate": start_date,
                        "EndDate": end_date,
                        "JobType": job_type
                    }
                    sales = pd.concat([sales, pd.DataFrame([new_row])], ignore_index=True)
                    
                    # 3) If there is an initial deposit, insert a matching Collections row
                    if deposit_paid_val and deposit_paid_val > 0:
                        collections = pd.concat([collections, pd.DataFrame([{
                            "QuoteID": new_id,
                            "Client": client,
                            "DepositDue": max(quoted_price_val - deposit_paid_val, 0.0),
                            "DepositPaid": deposit_paid_val,
                            "BalanceDue": max(quoted_price_val - deposit_paid_val, 0.0),
                            "Status": "Partially Paid" if deposit_paid_val < quoted_price_val else "Paid",
                        }])], ignore_index=True)

                    # 4) Recompute derived fields
                    sales = sync_deposit_paid(sales, collections)
                    collections = update_balance_due(sales, collections)

                    # 5) SAVE → only reload if save succeeded
                    if save_data(sales, collections, assignments):
                        sales, collections, assignments = load_data()
                        st.session_state["sale_added"] = new_row
                        st.rerun()
                    else:
                        st.stop()


    # Always show all sales at the bottom
    st.subheader("All Sales Entries")
    st.caption(f"Rows in Sales (after load): {len(sales)}")

    # Display-only formatting: $ for money, date-only, % for Deposit%
    display_sales = sales.copy()
    for money_col in ("QuotedPrice", "DepositPaid"):
        if money_col in display_sales.columns:
            display_sales[money_col] = _fmt_currency_series(pd.to_numeric(display_sales[money_col], errors="coerce"))
    if "Deposit%" in display_sales.columns:
        display_sales["Deposit%"] = _fmt_percent_series(pd.to_numeric(display_sales["Deposit%"], errors="coerce"))
    for date_col in ("StartDate", "EndDate"):
        if date_col in display_sales.columns and pd.api.types.is_datetime64_any_dtype(display_sales[date_col]):
            display_sales[date_col] = display_sales[date_col].dt.date
    st.dataframe(display_sales)

# ---------------- Collections ----------------

# ---------------- Collections ----------------
elif page == "Collections":
    st.title("💰 Collections")

    # Success banner after adding a collection
    if "collection_added_quote" in st.session_state and st.session_state["collection_added_quote"] is not None:
        st.success(f"✅ Collection successfully added for Quote ID {st.session_state['collection_added_quote']}")
        add_another = st.radio("Would you like to add another collection?", ["No", "Yes"], index=0)
        if add_another == "No":
            st.subheader("All Collections Data")
            collections_display = collections.drop(columns=["DepositDue"], errors="ignore").copy()
            for money_col in ("DepositPaid", "BalanceDue"):
                if money_col in collections_display.columns:
                    collections_display[money_col] = _fmt_currency_series(pd.to_numeric(collections_display[money_col], errors="coerce"))
            cols = [c for c in ["QuoteID","Client","DepositPaid","BalanceDue","Status"] if c in collections_display.columns]
            st.dataframe(collections_display[cols] if cols else collections_display)
            st.session_state["collection_added_quote"] = None
            st.stop()
        else:
            st.session_state["collection_added_quote"] = None
            st.session_state.pop("collection_submitted", None)

    # WON jobs only
    won_sales = sales[sales["Status"] == "Won"]
    options = ["Select a QuoteID or Client"] + [
        f"{int(row['QuoteID'])} - {row['Client']}" for _, row in won_sales.iterrows()
    ]
    selected = st.selectbox("Search QuoteID or Client", options, index=0)

    if selected != "Select a QuoteID or Client":
        selected_quote_id = int(selected.split(" - ")[0])

        # Base info
        sales_row = won_sales[won_sales["QuoteID"] == selected_quote_id]
        client_name = str(sales_row["Client"].values[0]) if not sales_row.empty else ""
        quoted_price = float(sales_row["QuotedPrice"].values[0]) if not sales_row.empty else 0.0

        # 👉 TOTAL paid to date comes straight from Sales (already rolled up by sync_deposit_paid)
        paid_to_date = float(
            sales.loc[sales["QuoteID"] == selected_quote_id, "DepositPaid"].values[0]
        ) if (sales["QuoteID"] == selected_quote_id).any() else 0.0

        # Collections history (raw ledger)
        coll_history = collections[collections["QuoteID"] == selected_quote_id].copy()
        coll_history["DepositPaid"] = pd.to_numeric(coll_history.get("DepositPaid", 0), errors="coerce").fillna(0.0)

        # Summary
        remaining_balance_due = max(quoted_price - paid_to_date, 0.0)
        pay_status = "Paid" if remaining_balance_due == 0 else ("Partially Paid" if paid_to_date > 0 else "Pending")

        st.markdown("### Quote Summary")
        c1, c2, c3 = st.columns(3)
        c1.metric("Quoted Price", f"${quoted_price:,.2f}")
        c2.metric("Paid To Date", f"${paid_to_date:,.2f}")
        c3.metric("Balance Due", f"${remaining_balance_due:,.2f}")
        st.caption(f"Collections ledger sum: ${coll_history['DepositPaid'].sum():,.2f} • Status: {pay_status}")

        # Sales info (display-only formatting for $, date-only, and %)
        st.markdown("#### Sales Info")
        if not sales_row.empty:
            sales_view_cols = [c for c in ["QuoteID","Client","QuotedPrice","Status","SalesRep","Deposit%","DepositPaid","StartDate","EndDate","JobType"] if c in sales_row.columns]
            display_sales_row = sales_row.copy()
            for money_col in ("QuotedPrice", "DepositPaid"):
                if money_col in display_sales_row.columns:
                    display_sales_row[money_col] = _fmt_currency_series(pd.to_numeric(display_sales_row[money_col], errors="coerce"))
            if "Deposit%" in display_sales_row.columns:
                display_sales_row["Deposit%"] = _fmt_percent_series(pd.to_numeric(display_sales_row["Deposit%"], errors="coerce"))
            for date_col in ("StartDate", "EndDate"):
                if date_col in display_sales_row.columns and pd.api.types.is_datetime64_any_dtype(display_sales_row[date_col]):
                    display_sales_row[date_col] = display_sales_row[date_col].dt.date
            st.dataframe(display_sales_row[sales_view_cols])
        else:
            st.info("No Sales row found for this Quote ID.")

        # Collections history table with running totals aligned to the Sales total
        st.markdown("#### Collections History")
        if not coll_history.empty:
            base_offset = max(0.0, paid_to_date - coll_history["DepositPaid"].sum())  # covers case where initial deposit isn't in Collections
            coll_history = coll_history.copy()
            coll_history["RunningTotal"] = coll_history["DepositPaid"].cumsum()
            coll_history["TotalPaidAfterThis"] = (base_offset + coll_history["RunningTotal"]).clip(upper=quoted_price)
            coll_history["BalanceAfterThis"] = (quoted_price - coll_history["TotalPaidAfterThis"]).clip(lower=0.0)

            cols_display = [c for c in ["QuoteID","Client","DepositPaid","Status","BalanceDue",
                                        "RunningTotal","TotalPaidAfterThis","BalanceAfterThis"]
                            if c in coll_history.columns or c in ["RunningTotal","TotalPaidAfterThis","BalanceAfterThis"]]
            # format money columns for display
            for money_col in ("DepositPaid", "BalanceDue", "RunningTotal", "TotalPaidAfterThis", "BalanceAfterThis"):
                if money_col in coll_history.columns:
                    coll_history[money_col] = _fmt_currency_series(pd.to_numeric(coll_history[money_col], errors="coerce"))
            st.dataframe(coll_history[cols_display])
        else:
            st.info("No collections for this Quote ID yet.")

        st.markdown("---")

        # Add new collection
        with st.form("update_collection"):
            st.write(f"Remaining Balance Due (before this collection): **${remaining_balance_due:,.2f}**")

            deposit_paid_input = st.number_input("New Collection Amount", value=0.0, min_value=0.0, format="%.2f")
            status_options = ["Pending", "Partially Paid", "Paid", "Overdue"]
            if not coll_history.empty and str(coll_history["Status"].iloc[-1]) in status_options:
                default_status_idx = status_options.index(str(coll_history["Status"].iloc[-1]))
            else:
                default_status_idx = 0
            status_input = st.selectbox("Status", options=status_options, index=default_status_idx)

            submitted = st.form_submit_button("Save Collection")

            if submitted and not st.session_state.get("collection_submitted", False):
                st.session_state["collection_submitted"] = True

                if deposit_paid_input <= 0:
                    st.error("⚠️ Please enter a collection amount greater than 0.00.")
                else:
                    new_row = {
                        "QuoteID": selected_quote_id,
                        "Client": client_name,
                        "DepositPaid": float(deposit_paid_input),
                        "Status": status_input,
                    }
                    collections = pd.concat([collections, pd.DataFrame([new_row])], ignore_index=True)

                    # Recompute + save (keeps Sales.DepositPaid = TOTAL; Collections.BalanceDue from Sales)
                    sales = sync_deposit_paid(sales, collections)
                    collections = update_balance_due(sales, collections)
                    if save_data(sales, collections, assignments):
                        st.session_state["collection_added_quote"] = selected_quote_id
                        safe_rerun()
                    else:
                        st.stop()

    # Always show all collections (without DepositDue)
    st.subheader("All Collections Data")
    collections_display = collections.drop(columns=["DepositDue"], errors="ignore").copy()
    for money_col in ("DepositPaid", "BalanceDue"):
        if money_col in collections_display.columns:
            collections_display[money_col] = _fmt_currency_series(pd.to_numeric(collections_display[money_col], errors="coerce"))
    cols = [c for c in ["QuoteID","Client","DepositPaid","BalanceDue","Status"] if c in collections_display.columns]
    st.dataframe(collections_display[cols] if cols else collections_display)


# ---------------- Assignments ----------------
elif page == "Assignments":
    st.title("📋 Assignments")

    # ---- Assigned and Pending Tasks KPI ----
    assigned_tasks = assignments["QuoteID"].unique() if "QuoteID" in assignments.columns else []
    won_sales = sales[sales["Status"] == "Won"] if "Status" in sales.columns else sales.iloc[0:0]
    pending_tasks = [quote_id for quote_id in won_sales["QuoteID"]] if "QuoteID" in sales.columns else []
    pending_tasks = [qid for qid in pending_tasks if qid not in assigned_tasks]

    col1, col2 = st.columns(2)
    col1.metric("Assigned Tasks", len(assigned_tasks))
    col2.metric("Pending Tasks", len(pending_tasks))

    # ---- Buttons to show Task Tables ----
    if st.button("Show Assigned Tasks"):
        assigned_task_data = assignments[assignments["QuoteID"].isin(assigned_tasks)] if "QuoteID" in assignments.columns else assignments
        st.subheader("Assigned Tasks")
        st.dataframe(assigned_task_data)

    if st.button("Show Pending Tasks"):
        pending_task_data = sales[sales["QuoteID"].isin(pending_tasks)] if "QuoteID" in sales.columns else sales.iloc[0:0]
        st.subheader("Pending Tasks")
        st.dataframe(pending_task_data)

    # Reload assignments (if needed)
    sales, collections, assignments = load_data()

    # ---- Assign Task Form ----
    st.subheader("Assign Tasks")
    won_sales = sales[sales["Status"] == "Won"] if "Status" in sales.columns else sales.iloc[0:0]

    if 'assigned' not in st.session_state:
        st.session_state.assigned = False

    if st.session_state.assigned:
        assign_another = st.radio("Do you want to assign another task?", options=["Yes", "No"])
        if assign_another == "No":
            st.write("You have chosen not to assign any more tasks.")
            st.session_state.assigned = False
        elif assign_another == "Yes":
            st.write("You can now assign another task.")
            st.session_state.assigned = False

    if not st.session_state.assigned:
        with st.form("assign_task"):
            options = ["Select Quote ID or Client"] + [
                f"{row['QuoteID']} - {row['Client']}" for _, row in won_sales.iterrows()
            ]
            selected_option = st.selectbox("Search QuoteID or Client", options, index=0)

            if selected_option != "Select Quote ID or Client" and "-" in selected_option:
                quote_id = int(selected_option.split(" - ")[0])
                client = selected_option.split(" - ")[1]
            else:
                quote_id = None
                client = None

            crew_member = st.text_input("Crew Member")
            start_date = st.date_input("Start Date")
            end_date = st.date_input("End Date")
            payment = st.number_input("Crew Payment", min_value=0.0, format="%.2f")
            notes = st.text_area("Notes")  # NEW: Notes input

            days_taken = (end_date - start_date).days if end_date and start_date else 0
            st.text(f"Days Taken: {days_taken} days")

            submitted = st.form_submit_button("Assign Task")

            if submitted:
                if not crew_member:
                    st.error("⚠️ Please enter the Crew Member name.")
                elif quote_id is None:
                    st.error("⚠️ Please select a valid Quote ID or Client.")
                else:
                    new_assignment = {
                        "QuoteID": quote_id,
                        "Client": client,
                        "CrewMember": crew_member,
                        "StartDate": start_date,
                        "EndDate": end_date,
                        "Payment": payment,
                        "DaysTaken": days_taken,
                        "Notes": notes,  # NEW: persist notes
                    }

                    assignments = pd.concat([assignments, pd.DataFrame([new_assignment])], ignore_index=True)
                    save_data(sales, collections, assignments)

                    st.success(f"Task assigned to {crew_member} for Quote ID {quote_id} successfully!")

                    sales, collections, assignments = load_data()
                    st.session_state.assigned = True
                    st.subheader("Updated Assignments")
                    # Hide Client column if present
                    assignments_display = assignments.drop(columns=["Client"]) if "Client" in assignments.columns else assignments
                    st.dataframe(assignments_display)
                    safe_rerun()

# ---------------- View Reports ----------------
if page == "View Reports":
    st.title("📊 View Reports")
    st.markdown("Here you can view and download reports for Sales Log, Collections, and Assignments.")

    sales, collections, assignments = load_data()

    report_tabs = st.radio("Select a Report to View", ("Sales Log", "Collections", "Assignments"))

    if report_tabs == "Sales Log":
        st.subheader("Sales Log")
        st.write(sales)
    elif report_tabs == "Collections":
        st.subheader("Collections")
        st.write(collections)
    elif report_tabs == "Assignments":
        st.subheader("Assignments")
        st.write(assignments)

    # Download button for unified report (Excel file with three tabs)
    st.subheader("Download Unified Report")
    st.markdown("Click the button below to download the report with Sales Log, Collections, and Assignments as separate tabs in an Excel file.")

    def create_excel_report(sales, collections, assignments):
        excel_file = io.BytesIO()
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            sales.to_excel(writer, sheet_name="Sales Log", index=False)
            collections.to_excel(writer, sheet_name="Collections", index=False)
            assignments.to_excel(writer, sheet_name="Assignments", index=False)
        excel_file.seek(0)
        return excel_file

    if st.button("Download Unified Report"):
        excel_file = create_excel_report(sales, collections, assignments)
        st.download_button(
            label="Download Unified Report (Excel)",
            data=excel_file,
            file_name="unified_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
